# coding=utf-8

from turtle import ontimer
import numpy as np
import pandas as pd
import openpyxl as xl

class DataFrameClass(pd.DataFrame):
    
    #-----------Basic properties------------------    
    # 数据转为二维数组
    @property
    def array(self):
        return np.asarray(self.values)
    
    # 数据转为矩阵
    @property
    def matrix(self):
        return np.mat(self.values)
        
    # 数据转为pandas 数据框
    @property
    def dataframe(self):
        return pd.DataFrame(self.values,columns=self.columns,index=self.index)
    
    # 属性：扩展矩阵时是否填充数据
    @property
    def amend_value(self):
        if not hasattr(self,'_amend_value'):
            self._amend_value=False
        return self._amend_value
    
    @amend_value.setter
    def amend_value(self, value):
        self._amend_value= value
    
    @property 
    def row_count(self):
        return self.shape[0]
    
    @property
    def col_count(self):
        return self.shape[1]
    
    @property
    def T(self):
        return DataFrameClass(super().T)
    
    
  #-----------Data Operation---------------------  
    def _size_ajust(self, other, r0, c0, value=0, value_row = -1, value_col=-1):       
        tmp = np.mat(other) 
        (row, col) = tmp.shape
        #(r0, c0) = self.shape 
        if row > r0:
            tmp = tmp[:r0,:]
        elif row < r0:
            if self.amend_value == True:
                v_rows  = np.ones((r0-row,col)) * value
            else:
                v_row = tmp[value_row,:]  
                v_rows = v_row.repeat(r0-row, axis=0) 
            tmp = np.vstack((tmp,v_rows))        
        (row, col) = tmp.shape         
        if col > c0:
            tmp = tmp[:,:c0]
        elif col < c0:
            if self.amend_value:
                v_cols  = np.ones((row,c0-col)) * value
            else:
                v_col = tmp[:,value_col]  
                v_cols = v_col.repeat(c0-col, axis=1)  
            tmp = np.hstack((tmp,v_cols))     
        return tmp 
    
    # 重载加法运算符
    def __add__(self, other):
        tmp = self._size_ajust(other, self.row_count, self.col_count)
        result = np.add(self.matrix, tmp)
        return DataFrameClass(result, columns=self.columns, index=self.index)
 
    def __sub__(self, other):
        tmp = self._size_ajust(other, self.row_count, self.col_count)
        result = np.subtract(self.matrix, tmp)
        return DataFrameClass(result, columns=self.columns, index=self.index)

    def __mul__(self, other):
        tmp = self._size_ajust(other, self.row_count, self.col_count, value=1)
        result = np.multiply(self.matrix, tmp)
        return DataFrameClass(result, columns=self.columns, index=self.index)

    def __truediv__(self, other):
        tmp = self._size_ajust(other, self.row_count, self.col_count, value=1)
        result = np.divide(self.matrix, tmp)
        return DataFrameClass(result, columns=self.columns, index=self.index)

    def copy(self,deep=True):
      if deep:
         tmp = pd.DataFrame.copy(self,deep)
         return DataFrameClass(tmp.values, columns=tmp.columns, index=tmp.index)
      else:
         return self

    def copy_head(self):
        tmp = DataFrameClass(columns=self.columns)
        return tmp

    def duplicate(self,times,axis=0):
        dflist = list([])
        for i in range(times):
            tmpDF = self.copy()
            dflist.append(tmpDF)
        df = pd.concat(dflist,axis=axis)
        return DataFrameClass(df.values,columns=df.columns,index=df.index)

    def subDataFrame(self,rstart=0,cstart=0,rend=-1,cend=-1):
        if rend==-1:
            rend = self.shape[0] - 1
        if cend==-1:
            cend = self.shape[1] - 1 
        tmp = self.iloc[rstart:rend+1, cstart:cend+1]   
        return DataFrameClass(tmp)

    def colDataFrame(self, cindex, rstart = 0, rend = -1):
        return self.subDataFrame(rstart=rstart, cstart=cindex, rend = rend, cend=cindex)        

    def colsDataFrame(self, cindexes, rstart = 0, rend = -1):
        tmp = DataFrameClass()
        columns = []
        for i in cindexes:
            sdf = self.colDataFrame(rstart=rstart, cindex=i, rend = rend)            
            columns.append(self.columns[i])
            tmp = tmp.concat(sdf)
        tmpo = DataFrameClass(tmp.values,columns=columns,index=tmp.index)
        return tmpo

    def rowDataFrame(self, rindex, cstart = 0, cend = -1):
        return self.subDataFrame(rstart=rindex, cstart=cstart, rend=rindex, cend=cend)

    def rowsDataFrame(self, rindexes, cstart = 0, cend = -1):
        tmp = DataFrameClass()
        index = []
        for i in rindexes:
            sdf = self.rowDataFrame(rindex=i, cstart=cstart, cend=cend)
            index.append(self.index[i])            
            tmp = tmp.append(sdf)
        tmpo = DataFrameClass(tmp.values, columns=tmp.columns, index=index)
        return tmpo

    def concat(self,sourceDFC, axis=1, inplace=False, ignore_index=False):
        #tmp = pd.concat([self.dataframe,sourceDFC.dataframe], axis=axis, ignore_index=ignore_index)
        tmp = pd.concat([self, sourceDFC], axis=axis, ignore_index=ignore_index)       
        tmp = DataFrameClass(tmp)
        if inplace:
            self.__init__(tmp)
        return tmp

    def clone(self, sourceDFC, columns=None, index=None):
        if isinstance(sourceDFC,pd.DataFrame) or isinstance(sourceDFC,DataFrameClass):
            self.__init__(sourceDFC.values,index=sourceDFC.index,columns=sourceDFC.columns)
        else:
            self.__init__(sourceDFC, columns=columns, index=index)
        return self

    def value_init(self,rows,cols,initValue=0,columns=None,index=None):
        tmp = np.ones((rows,cols)) * initValue
        self.__init__(tmp, columns=columns, index=index)
        return self
        
    def continuous_value_init(self,  rows, cols, v_start, v_interval=1,columns=None,index=None):
        tmp = np.arange(rows*cols) * v_interval + v_start
        tmp = tmp.reshape((rows,cols))
        self.__init__(tmp,columns=columns, index=index)      
        return self  
    
    def random_value_init(self,rows,cols,v_min=0, v_max=1, dtype=int, columns=None,index=None, round=3):
        if dtype==int:
            tmp = np.random.randint(v_min, v_max,(rows,cols))
        else:
            tmp = np.random.rand(rows,cols) * (v_max-v_min)  + v_min
            tmp = np.round(tmp,round)
        self.__init__(tmp,columns=columns, index=index)      
        return self        
              
    def normal_random_value_init(self,rows,cols, mean, std, columns=None,index=None, round=3):
        tmp = np.random.normal(mean, std, (rows,cols))
        tmp = np.round(tmp, round)
        self.__init__(tmp,columns=columns, index=index)      
        return self   
    
    def append(self,df, inplace=True):
        self.concat(df,axis=0, inplace=inplace)
        return self
    
    def resize(self, rows, cols, value_fill=0, value_row = -1, value_col=-1, columns=None, index=None, amend_value=True):
        a_v = self.amend_value
        self._amend_value = amend_value
        tmp_values = self.values
        tmp = self._size_ajust(tmp_values,rows,cols, value=value_fill,value_row=value_row,value_col=value_col)       
        self.__init__(tmp, columns=columns, index=index)
        self.amend_value = a_v
        return self     
    
    def drop_column(self, index, inplace=True):
        col_name = self.columns[index]
        self.drop(columns=[col_name],axis=1,inplace=inplace)
        return self
    
    def drop_row(self, index, inplace=True):
        row_index = self.index[index]
        self.drop(index=[row_index],axis=0,inplace=inplace)
        return self       

    def __reverse(self,axis=0,inplace=False):
        tmp = self.copy()
        maxrow = tmp.shape[0]
        maxcol = tmp.shape[1]
        if axis==0:
            for i in range(maxrow):
                tmp.drop_row(0,inplace=True)
                row = self.rowDataFrame(maxrow-i-1)
                tmp.append(row, inplace=True)
        else:
            for i in range(maxcol):
                tmp.drop_column(0,inplace=True)
                col = self.colDataFrame(maxcol-i-1)
                tmp.concat(col,inplace=True)
        if inplace:
            self.__init__(tmp)
        return tmp
    
    def reverse(self,axis=0,inplace=False):
        if axis==0:
            tmp = self.iloc[::-1,:]
        else:
            tmp = self.iloc[:,::-1]
        if inplace:
            self.__init__(tmp)
        return tmp

    def get_column_index(self, column):
        if not column in self.columns:
            index_column = -1
        else:
            for i in range(0,len(self.columns)):
                if column ==  self.columns[i]:
                    index_column = i
                    break
        return index_column


    def _compare(self,column,value,compare_method='equal'):
        result=False
        if compare_method == 'column_equal_value' or compare_method == 'equal' or compare_method == 'eq':
            result = (column == value)
        elif compare_method == 'column_include_value' or compare_method == 'include':
            result = (value in column)
        elif compare_method == 'column_in_value' or compare_method == 'in':
            result = (value in column)
        elif compare_method == 'column_greater_than_value' or compare_method == 'gt':
            result = (column > value)
        elif compare_method == 'column_greater_equal_than_value' or compare_method == 'ge':
            result = (column >= value)
        elif compare_method == 'column_less_than_value' or compare_method == 'lt':
            result = (column < value)
        elif compare_method == 'column_less_equal_than_value' or compare_method == 'le':
            result = (column <= value)

        return result
    def filter(self, column_name, value, compare_method='equal'):
        filter_data = self.copy_head()
        index_column = self.get_column_index(column_name)
        if index_column > 0 :
            for iRow in range(0,self.row_count):
                column_value = self.iloc[iRow,index_column]
                if self._compare(column_value,value,compare_method):
                    row_data = self.rowDataFrame(iRow)
                    filter_data.append(row_data)
        return filter_data






   #--------------Stat--------------------------------
   

class ExcelObject:
    def __init__(self):
        # 自定义类属性 book 用以存放 excel 数据表
        self._book = xl.Workbook()
        self._filename = ""

    def __del__(self):
        self._book.close()

    def new_workbook(self):
        self._book = xl.Workbook()
        self._filename = ""

    def load_workbook(self, path):
        self._filename = path
        self._book = xl.load_workbook(path)

    def save_workbook(self, path=None, newFilename=True):
        if path is None:
            self._book.save(self._filename)
        else:
            self._book.save(path)
            if newFilename:
                self._filename = path                
    
    @property
    def file_name(self):
        return self._filename
    
    @property
    def sheet_count(self):
        return len(self._book.worksheets)

    @property
    def sheet_names(self):
        return self._book.sheetnames
    
    @property
    def active_sheet(self):
        return self._book.active
    
    @active_sheet.setter
    def active_sheet(self, value):
        self._book.active = value  # self._book.worksheets[value]
        self.on_active_message()        
        
    @property
    def react_on_active_messagae(self):
        if not hasattr(self,'_react_on_active_messagae'):
            self._react_on_active_messagae = False
        return self._react_on_active_messagae
    
    @react_on_active_messagae.setter
    def react_on_active_messagae(self, value):
        self._react_on_active_messagae = value
        
    def on_active_message(self):
        if self.react_on_active_messagae:
            print('Current worksheet:' + self._book.active.title)


    def get_sheet(self, index):
        return self._book.worksheets[index]

    def get_sheet_by_name(self, sname):
        sheet_id = -1
        sheet_names = self.sheet_names
        for iSheet in range(0,self.sheet_count):
            if sname == sheet_names[iSheet]:
                sheet_id = iSheet
                break
        return sheet_id


    def is_sheet_exist(self, sname):
        return sname in self._book.sheetnames

    def change_sheet_name(self, sname, index):
        sheet = self._book.worksheets[index]
        sheet.title = sname

    def add_sheet(self, sname, index=None):
        # 添加新的表格
        sheet = self._book.create_sheet(sname, index)
        return sheet

    def delete_sheet_by_name(self, sname):
        # 删除表格
        ws = self._book[sname]
        self._book.remove(ws)

    def delete_sheet_by_index(self, index):
        ws = self._book.worksheets[index]
        self._book.remove(ws)

    def delete_all_sheets(self):
        for i in range(self.sheet_count):
            self.delete_sheet_by_index(0)

    def clear_sheet(self, index=None):
        if index is None:
            sheet = self._book.active
        else:
            sheet = self._book.worksheets[index]
        for row in sheet.rows:
            for cell in row:
                cell.value = ""

    def sheet_max_row(self, index=None):
        if index == None:
            sheet = self._book.active
        else:
            sheet = self._book.worksheets[index]
        return sheet.max_row

    def sheet_max_col(self, index=None):
        if index is None:
            sheet = self._book.active
        else:
            sheet = self._book.worksheets[index]
        return sheet.max_column

    def copy_sheet(self, datasheet, index=None, clc=True):
        if index is None:
            sheet = self._book.active
        else:
            sheet = self._book.worksheets[index]
        if clc:
            self.clear_sheet(index)
        sheet.title = datasheet.title
        mrow = datasheet.max_row + 1
        mcol = datasheet.max_column + 1
        for irow in range(1, mrow):
            for icol in range(1, mcol):
                sheet.cell(irow, icol).value = datasheet.cell(irow, icol).value

    def clone_sheet(self, datasheet, index=None):
        sheet = self.add_sheet(datasheet.title, index)
        if index is None:
            self._book.active = self.sheet_count - 1
        else:
            self._book.active = index
        self.copy_sheet(datasheet)

    def cell_values(self, rstart=2, cstart=2, rend=0, cend=0, index=None):
        if index is None:
            sheet = self.active_sheet
        else:
            sheet = self._book.worksheets[index]
        if rend == 0:
            rend = sheet.max_row
        if cend == 0:
            cend = sheet.max_column
        cnum = cend - cstart + 1
        rnum = rend - rstart + 1
        values = []
        for i in range(rnum):
            values.append([None] * cnum)
        valDF = pd.DataFrame(values)
        for r in range(rstart, rend + 1):
            for c in range(cstart, cend + 1):
                valDF[c - cstart][r - rstart] = sheet.cell(r, c).value
        return valDF

    def get_columns(self, index=None, cstart=1, cend=0):
        if index == None:
            sheet = self.active_sheet
        else:
            sheet = self._book.worksheets[index]
        if cend == 0:
            cend = sheet.max_column
        cols = []
        for i in range(cstart, cend + 1):
            cols.append(sheet.cell(1, i).value)
        return cols

    def get_index(self, index=None, rstart=2, rend=0):
        if index is None:
            sheet = self.active_sheet
        else:
            sheet = self._book.worksheets[index]
        if rend == 0:
            rend = sheet.max_row
        rows = []
        for i in range(rstart, rend + 1):
            rows.append(sheet.cell(i, 1).value)
        return rows

    def cell_DataFrame(self, rstart=2, cstart=2, rend=0, cend=0, index=None, withTitles=True, withIndex=True):
        data = self.cell_values(rstart, cstart, rend, cend, index)
        if withTitles:
            data.columns = self.get_columns(index, cstart, cend)
        if withIndex:
            data.index = self.get_index(index, rstart, rend)
        return DataFrameClass(data)

    def cell_matrix(self, rstart=2, cstart=2, rend=0, cend=0, index=None):
        cm = self.cell_values(rstart, cstart, rend, cend, index)
        return np.mat(cm)

    def cell_col_list(self,cstart,rstart=2,rend=0,index=None):
        if index == None:
            sheet = self.active_sheet
        else:
            sheet = self._book.worksheets[index]
        if rend == 0:
            rend = sheet.max_row
        cols = []
        for i in range(rstart, rend + 1):
            cols.append(sheet.cell(i,cstart).value)
        return cols
    
    def copy_list(self, sourcelist, horizontal=True, rstart=2, cstart=2, index=None):
        if index != None:
            self.active_sheet= index
        litems = len(sourcelist)
        if horizontal:
            for i in range(litems):
                self._book.active.cell(rstart, cstart + i).value = sourcelist[i]
        else:
            for i in range(litems):
                self._book.active.cell(rstart + i, cstart).value = sourcelist[i]

    def copy_array(self, sourcearray, rstart=2, cstart=2, index=None):
        if index is None:
            sheet = self.active_sheet
        else:
            sheet = self._book.worksheets[index]
        rows = len(sourcearray)
        cols = len(sourcearray[0])
        for i in range(rstart, rstart + rows):
            for j in range(cstart, cstart + cols):
                sheet.cell(i, j).value = sourcearray[i - rstart][j - cstart]

    def copy_matrix(self, sourcematrix, rstart=2, cstart=2, index=None):
        if index is None:
            sheet = self.active_sheet
        else:
            sheet = self._book.worksheets[index]
        rows = sourcematrix.shape[0]
        cols = sourcematrix.shape[1]
        for i in range(rstart, rstart + rows):
            for j in range(cstart, cstart + cols):
                sheet.cell(i, j).value = sourcematrix[i - rstart, j - cstart]

    def copy_DataFrame(self, sourcedataframe, rstart=2, cstart=2, index=None, withtitles=True, withindex=True,
                      indextitle='ID'):
        if index is None:
            sheet = self.active_sheet
        else:
            sheet = self._book.worksheets[index]
        rows = sourcedataframe.shape[0]
        cols = sourcedataframe.shape[1]
        for i in range(rstart, rstart + rows):
            for j in range(cstart, cstart + cols):
                sheet.cell(i, j).value = sourcedataframe.iloc[i - rstart, j - cstart]
        if withtitles:
            self.copy_list(sourcedataframe.columns, rstart=1, cstart=cstart, index=index)
        if withindex:
            self.copy_list(sourcedataframe.index, horizontal=False, rstart=rstart, cstart=1, index=index)
            sheet.cell(1, 1).value = indextitle

    def variable_from_sheets(self,var_index, column_names = None, index_names=None,withtitles=True, withindex=True):
        data_name = self.cell_values(rstart=1,rend=1,cstart=var_index,cend=var_index)[0][0]
        data_index = self.sheet_names
        data_columns = []
        rstart = 2
        if withindex:
            if withtitles:
                data_columns = self.get_index(rstart=rstart)
            else:
                rstart = 1
                data_columns = self.get_index(rstart=rstart)
        else:
            data_columns = column_names
        data = pd.DataFrame(columns=data_columns)
        for i in range(self.sheet_count):
            tmp_data = self.cell_col_list(rstart=rstart,cstart=var_index,index=i)
            row = pd.DataFrame([tmp_data],columns=data_columns,\
                               index=[data_index[i]])
            data = data.append(row)
        return [data_name,data]

 

   