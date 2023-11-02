# coding=utf-8
import openpyxl as xl
import pandas as pd
import numpy as np
from itertools import chain

class ExcelHelper:

    def __init__(self):
        # 自定义类属性 book 用以存放 excel 数据表
        self._book = xl.Workbook()
        self._filename = ""

    def __del__(self):
        self._book.close()

    def load_workbook(self, path):
        self._filename = path
        self._book = xl.load_workbook(path)

    def save_workbook(self, path=None, newFilename=True):
        if path is None:
            self._book.save(self._filename)
        else:
            self._book.save(path)
            if newFilename:
                self._filename = path

    def file_name(self):
        return self._filename

    def sheet_count(self):
        return len(self._book.worksheets)

    def sheet_names(self):
        return self._book.sheetnames

    def active_sheet(self):
        return self._book.active

    def set_active_sheet(self, value):
        self._book.active = value  # self._book.worksheets[value]
        print('Current worksheet:' + self._book.active.title)

    def get_sheet(self, index):
        return self._book.worksheets[index]

    def is_sheet_exist(self, sname):
        return sname in self._book.sheetnames

    def change_sheet_name(self, sname, index):
        sheet = self._book.worksheets[index]
        sheet.title = sname

    def add_sheet(self, sname, index=None):
        # 添加新的表格
        sheet = self._book.create_sheet(sname, index)
        return sheet

    def delete_sheet_by_name(self, sname):
        # 删除表格
        ws = self._book[sname]
        self._book.remove(ws)

    def delete_sheet_by_index(self, index):
        ws = self._book.worksheets[index]
        self._book.remove(ws)

    def delete_all_sheets(self):
        for i in range(self.sheet_count()):
            self.delete_sheet_by_index(0)

    def clear_sheet(self, index=None):
        if index is None:
            sheet = self._book.active
        else:
            sheet = self._book.worksheets[index]
        for row in sheet.rows:
            for cell in row:
                cell.value = ""

    def sheet_max_row(self, index=None):
        if index == None:
            sheet = self._book.active
        else:
            sheet = self._book.worksheets[index]
        return sheet.max_row

    def sheet_max_col(self, index=None):
        if index is None:
            sheet = self._book.active
        else:
            sheet = self._book.worksheets[index]
        return sheet.max_column

    def copy_sheet(self, datasheet, index=None, clc=True):
        if index is None:
            sheet = self._book.active
        else:
            sheet = self._book.worksheets[index]
        if clc:
            self.clear_sheet(index)
        sheet.title = datasheet.title
        mrow = datasheet.max_row + 1
        mcol = datasheet.max_column + 1
        for irow in range(1, mrow):
            for icol in range(1, mcol):
                sheet.cell(irow, icol).value = datasheet.cell(irow, icol).value

    def clone_sheet(self, datasheet, index=None):
        sheet = self.add_sheet(datasheet.title, index)
        if index is None:
            self._book.active = self.sheet_count() - 1
        else:
            self._book.active = index
        self.copy_sheet(datasheet)

    def cell_values(self, rstart=2, cstart=2, rend=0, cend=0, index=None):
        if index is None:
            sheet = self.active_sheet()
        else:
            sheet = self._book.worksheets[index]
        if rend == 0:
            rend = sheet.max_row
        if cend == 0:
            cend = sheet.max_column
        cnum = cend - cstart + 1
        rnum = rend - rstart + 1
        values = []
        for i in range(rnum):
            values.append([None] * cnum)
        valDF = pd.DataFrame(values)
        for r in range(rstart, rend + 1):
            for c in range(cstart, cend + 1):
                valDF[c - cstart][r - rstart] = sheet.cell(r, c).value
        return valDF

    def get_columns(self, index=None, cstart=1, cend=0):
        if index == None:
            sheet = self.active_sheet()
        else:
            sheet = self._book.worksheets[index]
        if cend == 0:
            cend = sheet.max_column
        cols = []
        for i in range(cstart, cend + 1):
            cols.append(sheet.cell(1, i).value)
        return cols

    def get_index(self, index=None, rstart=2, rend=0):
        if index is None:
            sheet = self.active_sheet()
        else:
            sheet = self._book.worksheets[index]
        if rend == 0:
            rend = sheet.max_row
        rows = []
        for i in range(rstart, rend + 1):
            rows.append(sheet.cell(i, 1).value)
        return rows

    def cell_DataFrame(self, rstart=2, cstart=2, rend=0, cend=0, index=None, withTitles=True, withIndex=True):
        data = self.cell_values(rstart, cstart, rend, cend, index)
        if withTitles:
            data.columns = self.get_columns(index, cstart, cend)
        if withIndex:
            data.index = self.get_index(index, rstart, rend)
        return data

    def cell_matrix(self, rstart=2, cstart=2, rend=0, cend=0, index=None):
        cm = self.cell_values(rstart, cstart, rend, cend, index)
        return np.mat(cm)

    def cell_col_list(self,cstart,rstart=2,rend=0,index=None):
        if index == None:
            sheet = self.active_sheet()
        else:
            sheet = self._book.worksheets[index]
        if rend == 0:
            rend = sheet.max_row
        cols = []
        for i in range(rstart, rend + 1):
            cols.append(sheet.cell(i,cstart).value)
        return cols


    def copy_list(self, sourcelist, horizontal=True, rstart=2, cstart=2, index=None):
        if index != None:
            self.set_active_sheet(index)
        litems = len(sourcelist)
        if horizontal:
            for i in range(litems):
                self._book.active.cell(rstart, cstart + i).value = sourcelist[i]
        else:
            for i in range(litems):
                self._book.active.cell(rstart + i, cstart).value = sourcelist[i]

    def copy_array(self, sourcearray, rstart=2, cstart=2, index=None):
        if index is None:
            sheet = self.active_sheet()
        else:
            sheet = self._book.worksheets[index]
        rows = len(sourcearray)
        cols = len(sourcearray[0])
        for i in range(rstart, rstart + rows):
            for j in range(cstart, cstart + cols):
                sheet.cell(i, j).value = sourcearray[i - rstart][j - cstart]

    def copy_matrix(self, sourcematrix, rstart=2, cstart=2, index=None):
        if index is None:
            sheet = self.active_sheet()
        else:
            sheet = self._book.worksheets[index]
        rows = sourcematrix.shape[0]
        cols = sourcematrix.shape[1]
        for i in range(rstart, rstart + rows):
            for j in range(cstart, cstart + cols):
                sheet.cell(i, j).value = sourcematrix[i - rstart, j - cstart]

    def copy_DataFrame(self, sourcedataframe, rstart=2, cstart=2, index=None, withtitles=False, withindex=False,
                      indextitle=''):
        if index is None:
            sheet = self.active_sheet()
        else:
            sheet = self._book.worksheets[index]
        rows = sourcedataframe.shape[0]
        cols = sourcedataframe.shape[1]
        for i in range(rstart, rstart + rows):
            for j in range(cstart, cstart + cols):
                sheet.cell(i, j).value = sourcedataframe.iloc[i - rstart, j - cstart]
        if withtitles:
            self.copy_list(sourcedataframe.columns, rstart=1, cstart=cstart, index=index)
        if withindex:
            self.copy_list(sourcedataframe.index, horizontal=False, rstart=rstart, cstart=1, index=index)
            sheet.cell(1, 1).value = indextitle

    def variable_from_sheets(self,var_index, column_names = None, index_names=None,withtitles=True, withindex=True):
        data_name = self.cell_values(rstart=1,rend=1,cstart=var_index,cend=var_index)[0][0]
        data_index = self.sheet_names()
        data_columns = []
        rstart = 2

        if withindex:
            if withtitles:
                data_columns = self.get_index(rstart=rstart)
            else:
                rstart = 1
                data_columns = self.get_index(rstart=rstart)
        else:
            data_columns = column_names

        data = pd.DataFrame(columns=data_columns)

        for i in range(self.sheet_count()):
            tmp_data = self.cell_col_list(rstart=rstart,cstart=var_index,index=i)
            row = pd.DataFrame([tmp_data],columns=data_columns,\
                               index=[data_index[i]])
            data = data.append(row)


        return [data_name,data]


